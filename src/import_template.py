import os
import traceback

from openpyxl import load_workbook
from PySide6.QtWidgets import QFileDialog, QLabel, QMessageBox, QTextEdit

from cache import Cache
from .clear_fields import clear_fields


def get_target_rows(active_sheet) -> list[int]:
        """Identify the rows of the spreadsheet meant to receive the PGC data"""
        target_rows:list[int] = []

        row_iterator:int = 1
        last_row:int = active_sheet.max_row
        
        while row_iterator <= last_row:
            module_id_cell_value:str | None = active_sheet.cell(row = row_iterator, column = 1).value
            module_input_cell_value:str | None = active_sheet.cell(row = row_iterator, column = 4).value

            # If column A of the row is filled and column D is empty, consider it a target row
            if module_id_cell_value != None and module_input_cell_value == None:
                target_rows.append(row_iterator)
            
            row_iterator += 1

        return target_rows


def get_module_info(target_rows:list[int], active_sheet) -> list[str]:
    """Retrieve the name and identifiers of the target rows"""
    module_info:list[str] = []

    for target_row in target_rows:
        module_info.append('{} {} {} (R{})'.format(active_sheet.cell(row = target_row, column = 1).value,
                                                   active_sheet.cell(row = target_row, column = 2).value,
                                                   active_sheet.cell(row = target_row, column = 3).value,
                                                   target_row))
    
    return module_info


def add_module_fields(module_info:list[str], main_window) -> None:
    """Add the entry fields and associated labels to the fields widget"""
    target_row_index:int = 0

    while target_row_index < len(module_info):
        content_label = QLabel('{}:'.format(module_info[target_row_index]))
        main_window.central_widget.scroll_area.inner_widget.fields_widget.fields_widget_layout.addWidget(content_label)

        content_textedit = QTextEdit()
        content_textedit.setAcceptRichText(False)
        main_window.central_widget.scroll_area.inner_widget.fields_widget.fields_widget_layout.addWidget(content_textedit)

        content_label = QLabel('')
        main_window.central_widget.scroll_area.inner_widget.fields_widget.fields_widget_layout.addWidget(content_label)

        target_row_index += 1


def import_template(toolbar, main_window) -> None:
    """Import the PGC template spreadsheet"""
    root_directory:str = os.path.abspath(os.sep)

    template_workbook_path:str = QFileDialog.getOpenFileName(toolbar, 'Select template spreadsheet', root_directory, 'Excel 2007-365 (*.xlsx)')[0]

    if not template_workbook_path:
        return

    Cache().write('template_workbook_path', template_workbook_path)

    try:
        # Disable the import button of the toolbar
        toolbar.import_template_action.setDisabled(True)

        template_workbook = load_workbook(template_workbook_path)

        if len(template_workbook.sheetnames) > 1:
            raise Exception('several sheets were detected in the provided template spreadsheet (expected only one)')
        
        template_sheet = template_workbook.active

        target_rows:list[int] = get_target_rows(template_sheet)

        if len(target_rows) == 0:
            raise Exception('no empty PGC modules were found in the provided template spreadsheet')
        
        Cache().write('target_rows', target_rows)

        clear_fields(main_window, False)

        module_info:list[str] = get_module_info(target_rows, template_sheet)
        add_module_fields(module_info, main_window)

        # Enable the create button of the buttons widget
        main_window.central_widget.scroll_area.inner_widget.buttons_widget.create_button.setDisabled(False)
        # Enable the import button of the toolbar
        toolbar.import_template_action.setDisabled(False)
    except Exception as e:
        toolbar.import_template_action.setDisabled(False)

        traceback_message = traceback.format_exc()
        
        QMessageBox.critical(toolbar,
                'An error occurred',
                'Error: {}\n\n{}'.format(e, traceback_message),
                QMessageBox.Ok)