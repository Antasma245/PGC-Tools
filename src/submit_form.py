import os
from re import sub
from time import time
import traceback

from openpyxl import load_workbook
from PySide6.QtWidgets import QMessageBox, QTextEdit

from cache import Cache
from .clear_fields import clear_fields
from settings import Settings


def tsv_to_list(tsv_string:str) -> list[list[str]]:
    """Convert a TSV-formatted string into a two dimensional list"""
    lines:list[str] = tsv_string.split('\n')

    tsv_list:list[list[str]] = []
    
    for line in lines:
        values:list[str] = line.split('\t')
        tsv_list.append(values)
    
    return tsv_list


def create_batch_sheet(template_workbook_path:str, target_rows:list[int], form_data:list[str]) -> str:
    """Create the PGC spreadsheet based on the submitted texts and template"""
    template_workbook = load_workbook(template_workbook_path)
    template_sheet = template_workbook.active

    content_list:list[list[list[str]]] = []

    for data_block in form_data:
        # Remove the eventual \n character added at the end of the string if the texts were pasted from LibreOffice Calc
        data_block:str = sub('\\n(?!.)', '', data_block)
        content_list.append(tsv_to_list(data_block))
    
    out_dir:str = Settings().read('out_dir')

    if not os.path.exists(out_dir):
        os.makedirs(out_dir)
    
    out_path:str = '{}/PGC_BATCH_OUT_{}.xlsx'.format(out_dir, int(time()))

    content_amount:int = len(content_list[0])

    language_amount:int = Settings().read('language_amount')

    for processed_sheet in range(content_amount):
        # Duplicate the template sheet
        duplicated_sheet = template_workbook.copy_worksheet(template_sheet)

        # For each target row
        for row_index in range(len(target_rows)):
            content_line:list[str] = content_list[row_index][processed_sheet]

            # If the line of content only has one entity (i.e. language), consider it a link
            if len(content_line) == 1:
                link:str = content_line[0]

                # Duplicate the link of the line of content to have one per language
                for _ in range(language_amount - 1):
                    content_line.append(link)
            
            # Iterate through the target range
            for target_range in duplicated_sheet.iter_rows(min_col = 4,
                                                    min_row = target_rows[row_index],
                                                    max_col = language_amount + 3,
                                                    max_row = target_rows[row_index]):
                cell_iterator:int = 0

                # For each cell of the target range
                for cell in target_range:
                    cell.value = content_line[cell_iterator]
                    cell_iterator += 1

        duplicated_sheet.title = str(processed_sheet + 1)

    # Unnecessary, just to improve the clarity of the subsequent operations
    out_workbook = template_workbook
    
    # Delete the template sheet from the generated spreadsheet
    del out_workbook[out_workbook.sheetnames[0]]

    out_workbook.save(out_path)

    return out_path


def submit_form(main_window) -> None:
    """Retrieve and process the texts from the fields of the fields widget"""
    main_window.central_widget.scroll_area.inner_widget.buttons_widget.create_button.setDisabled(True)

    template_workbook_path:str = Cache().read('template_workbook_path')
    target_rows:list[int] = Cache().read('target_rows')

    form_data:list[str] = []

    fields_widget = main_window.central_widget.scroll_area.inner_widget.fields_widget

    # Retrieve the texts from the fields of the fields widget
    for field in fields_widget.findChildren(QTextEdit):
        form_data.append(field.toPlainText())

    try:
        out_path = create_batch_sheet(template_workbook_path, target_rows, form_data)

        QMessageBox.information(fields_widget,
                                'Process successfull',
                                'Your PGC spreadsheet has been generated and saved to {}'.format(out_path),
                                QMessageBox.Ok)

        clear_fields(main_window, True)
    except Exception as e:
        main_window.central_widget.scroll_area.inner_widget.buttons_widget.create_button.setDisabled(False)

        traceback_message = traceback.format_exc()

        QMessageBox.critical(fields_widget,
                        'An error occured',
                        'Error: {}\n\n{}'.format(e, traceback_message),
                        QMessageBox.Ok)